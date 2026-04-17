"""
excel_tracker.py  (v2 – Premium Dashboard Edition)
────────────────────────────────────────────────────
Professioneller Trading-Tracker mit:
  • Dashboard-Sheet mit KPI-Cards und eingebetteten Charts
  • Modernes Dark-Theme mit konsistenter Design-Sprache
  • Conditional Formatting für P&L, Win-Rate, Konfidenz
  • Equity Curve (Line Chart) + Monthly P&L (Bar Chart) + Asset-Performance (Pie)
  • Automatische Kapital-Berechnung über Excel-Formeln
  • Data Bars für visuelle Performance-Anzeige
"""

import os
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, CellIsRule, FormulaRule, Rule
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.drawing.colors import ColorChoice as DrawColorChoice

log = logging.getLogger(__name__)

# ─── Premium Color Palette (Bloomberg / TradingView Inspired) ────────────────
class Colors:
    # Backgrounds (Dark, refined)
    BG_DEEP      = "FF0A0E1A"   # Page background
    BG_SURFACE   = "FF141823"   # Card background
    BG_ELEVATED  = "FF1C2130"   # Elevated card
    BG_HEADER    = "FF0F1420"   # Header background
    BG_ACCENT    = "FF1E2332"   # Zebra stripe

    # Borders
    BORDER_SOFT  = "FF1F2937"
    BORDER_MID   = "FF2D3748"
    BORDER_GOLD  = "FFD4AF37"

    # Accent colors
    GOLD         = "FFD4AF37"   # Refined gold
    GOLD_LIGHT   = "FFE8C86A"
    CYAN         = "FF06B6D4"
    CYAN_LIGHT   = "FF67E8F9"
    EMERALD      = "FF10B981"
    EMERALD_LIGHT= "FF6EE7B7"
    RED          = "FFEF4444"
    RED_LIGHT    = "FFFCA5A5"
    PURPLE       = "FF8B5CF6"
    PURPLE_LIGHT = "FFC4B5FD"
    AMBER        = "FFF59E0B"
    BLUE         = "FF3B82F6"

    # Text
    TEXT_PRIMARY   = "FFE5E7EB"
    TEXT_SECONDARY = "FF9CA3AF"
    TEXT_MUTED     = "FF6B7280"
    WHITE          = "FFFFFFFF"


# ─── Helper Factories ─────────────────────────────────────────────────────────
def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def font(bold=False, color=Colors.TEXT_PRIMARY, size=11, name="Segoe UI"):
    return Font(name=name, bold=bold, color=color, size=size)


def thin_border(color=Colors.BORDER_SOFT):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def bottom_accent_border(color, weight="medium"):
    return Border(
        bottom=Side(border_style=weight, color=color),
        left=Side(border_style="thin", color=Colors.BORDER_SOFT),
        right=Side(border_style="thin", color=Colors.BORDER_SOFT),
    )


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def right():
    return Alignment(horizontal="right", vertical="center", indent=1)


# ─── Main Tracker Class ──────────────────────────────────────────────────────
class ExcelTracker:
    """Premium Trading Tracker with dashboard, charts, and conditional formatting."""

    SHEETS = ["Dashboard", "Demo-Kapital", "Trades", "Analyse-Log", "Performance", "Einstellungen"]

    def __init__(self, data_dir: str):
        self.path = Path(data_dir) / "Trading_Tracker.xlsx"
        self._init_workbook()

    def _init_workbook(self):
        if self.path.exists():
            log.info(f"Excel geladen: {self.path}")
            return
        log.info(f"Excel wird erstellt: {self.path}")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Daten-Sheets zuerst (Dashboard referenziert diese)
        self._create_demo_kapital_sheet(wb)
        self._create_trades_sheet(wb)
        self._create_analyse_sheet(wb)
        self._create_performance_sheet(wb)
        self._create_settings_sheet(wb)
        # Dashboard zuletzt
        self._create_dashboard_sheet(wb)
        # Dashboard als erstes Sheet in der Reihenfolge einordnen
        dashboard = wb["Dashboard"]
        wb.move_sheet(dashboard, offset=-(len(wb.sheetnames) - 1))
        wb.active = 0
        wb.save(self.path)
        log.info("✅ Excel erstellt")

    # ─── Sheet: Dashboard (NEW – Premium hero page) ──────────────────────────
    def _create_dashboard_sheet(self, wb):
        ws = wb.create_sheet("Dashboard")
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 110
        ws.sheet_properties.tabColor = Colors.GOLD.replace("FF", "", 1)

        # Kompakteres Layout (13 Cols statt 18)
        ws.column_dimensions["A"].width = 2  # left margin
        for col in range(2, 14):
            ws.column_dimensions[get_column_letter(col)].width = 13
        ws.column_dimensions["N"].width = 2  # right margin

        # Full page background
        self._paint_background(ws, rows=48, cols=14, color=Colors.BG_DEEP)

        # ── Hero Header ──
        ws.merge_cells("B2:M2")
        h = ws["B2"]
        h.value = "TRADING COMMAND CENTER"
        h.font = Font(name="Segoe UI", bold=True, color=Colors.GOLD, size=20)
        h.fill = fill(Colors.BG_HEADER)
        h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[2].height = 40

        ws.merge_cells("B3:M3")
        sub = ws["B3"]
        sub.value = "Premium Multi-Agent Trading Dashboard  •  Live-Daten aus Trades & Demo-Kapital"
        sub.font = font(color=Colors.TEXT_SECONDARY, size=10)
        sub.fill = fill(Colors.BG_HEADER)
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[3].height = 22

        ws.row_dimensions[4].height = 8

        # ── KPI Cards Row (B5:M9) – 4 cards, 3 cols each ──
        self._kpi_card(ws, "B5:D9", "AKTUELLES KAPITAL",
                       "='Demo-Kapital'!D4", "€#,##0.00",
                       Colors.CYAN, icon="◆")
        self._kpi_card(ws, "E5:G9", "GESAMT P&L",
                       "='Demo-Kapital'!F4", "€#,##0.00;[Red]-€#,##0.00",
                       Colors.GOLD, icon="▲")
        self._kpi_card(ws, "H5:J9", "ROI",
                       "=IF('Demo-Kapital'!B4>0,('Demo-Kapital'!D4-'Demo-Kapital'!B4)/'Demo-Kapital'!B4,0)",
                       "0.00%;[Red]-0.00%",
                       Colors.PURPLE, icon="%")
        self._kpi_card(ws, "K5:M9", "WIN RATE",
                       '=IF(COUNTA(Trades!A2:A10000)>0,COUNTIF(Trades!J2:J10000,"gewonnen")/COUNTA(Trades!A2:A10000),0)',
                       "0.0%",
                       Colors.EMERALD, icon="★")

        ws.row_dimensions[10].height = 8

        # ── Second row: smaller stat tiles ──
        self._stat_tile(ws, "B11:D13", "GESAMT TRADES",
                        "=COUNTA(Trades!A2:A10000)", "0",
                        Colors.BLUE)
        self._stat_tile(ws, "E11:G13", "GEWONNEN",
                        '=COUNTIF(Trades!J2:J10000,"gewonnen")', "0",
                        Colors.EMERALD)
        self._stat_tile(ws, "H11:J13", "VERLOREN",
                        '=COUNTIF(Trades!J2:J10000,"verloren")', "0",
                        Colors.RED)
        self._stat_tile(ws, "K11:M13", "BESTER TRADE",
                        "=IFERROR(MAX(Trades!K2:K10000),0)", "€#,##0.00",
                        Colors.GOLD)

        ws.row_dimensions[14].height = 8

        # ── Equity curve section header ──
        ws.merge_cells("B15:M15")
        s1 = ws["B15"]
        s1.value = "   EQUITY CURVE  —  Kapitalverlauf über Zeit"
        s1.font = Font(name="Segoe UI", bold=True, color=Colors.CYAN, size=11)
        s1.fill = fill(Colors.BG_ELEVATED)
        s1.alignment = Alignment(horizontal="left", vertical="center")
        s1.border = bottom_accent_border(Colors.CYAN, "medium")
        ws.row_dimensions[15].height = 24

        for r in range(16, 32):
            ws.row_dimensions[r].height = 15

        # ── Bottom section: pie + bar charts ──
        ws.merge_cells("B33:G33")
        ws["B33"].value = "   WIN / LOSS VERTEILUNG"
        ws["B33"].font = Font(name="Segoe UI", bold=True, color=Colors.EMERALD, size=11)
        ws["B33"].fill = fill(Colors.BG_ELEVATED)
        ws["B33"].alignment = Alignment(horizontal="left", vertical="center")
        ws["B33"].border = bottom_accent_border(Colors.EMERALD, "medium")

        ws.merge_cells("H33:M33")
        ws["H33"].value = "   ASSET-PERFORMANCE"
        ws["H33"].font = Font(name="Segoe UI", bold=True, color=Colors.PURPLE, size=11)
        ws["H33"].fill = fill(Colors.BG_ELEVATED)
        ws["H33"].alignment = Alignment(horizontal="left", vertical="center")
        ws["H33"].border = bottom_accent_border(Colors.PURPLE, "medium")
        ws.row_dimensions[33].height = 24

        for r in range(34, 48):
            ws.row_dimensions[r].height = 15

        # Render charts
        self._render_dashboard_charts(ws)

        # Footer
        ws.merge_cells("B49:M49")
        ws["B49"].value = f"  Generiert: {datetime.now().strftime('%d.%m.%Y %H:%M')}  •  KPIs und Charts werden automatisch aus den Datenblättern aktualisiert"
        ws["B49"].font = font(color=Colors.TEXT_MUTED, size=9)
        ws["B49"].fill = fill(Colors.BG_DEEP)
        ws["B49"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[49].height = 20

    def _kpi_card(self, ws, range_ref, label, value_formula, number_format, accent_color, icon=""):
        """Render a premium KPI card with label + large value."""
        start, end = range_ref.split(":")
        start_col = openpyxl.utils.column_index_from_string(''.join(c for c in start if c.isalpha()))
        start_row = int(''.join(c for c in start if c.isdigit()))
        end_col = openpyxl.utils.column_index_from_string(''.join(c for c in end if c.isalpha()))
        end_row = int(''.join(c for c in end if c.isdigit()))

        mid_row = start_row + 1

        label_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{mid_row}"
        value_range = f"{get_column_letter(start_col)}{mid_row+1}:{get_column_letter(end_col)}{end_row}"

        ws.merge_cells(label_range)
        ws.merge_cells(value_range)

        # Label
        lbl = ws[f"{get_column_letter(start_col)}{start_row}"]
        lbl.value = f"  {icon}  {label}"
        lbl.font = Font(name="Segoe UI", bold=True, color=Colors.TEXT_SECONDARY, size=9)
        lbl.fill = fill(Colors.BG_ELEVATED)
        lbl.alignment = Alignment(horizontal="left", vertical="center")

        # Value
        val = ws[f"{get_column_letter(start_col)}{mid_row+1}"]
        val.value = value_formula
        val.font = Font(name="Segoe UI", bold=True, color=accent_color, size=22)
        val.fill = fill(Colors.BG_ELEVATED)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.number_format = number_format

        # Borders around the card
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill(Colors.BG_ELEVATED)
                # Outer border
                left_b = Side(border_style="thin", color=Colors.BORDER_SOFT) if c == start_col else None
                right_b = Side(border_style="thin", color=Colors.BORDER_SOFT) if c == end_col else None
                top_b = Side(border_style="thin", color=Colors.BORDER_SOFT) if r == start_row else None
                bot_b = Side(border_style="medium", color=accent_color) if r == end_row else None
                if any([left_b, right_b, top_b, bot_b]):
                    cell.border = Border(left=left_b, right=right_b, top=top_b, bottom=bot_b)

    def _stat_tile(self, ws, range_ref, label, value_formula, number_format, accent_color):
        """Smaller stat tile."""
        start, end = range_ref.split(":")
        start_col = openpyxl.utils.column_index_from_string(''.join(c for c in start if c.isalpha()))
        start_row = int(''.join(c for c in start if c.isdigit()))
        end_col = openpyxl.utils.column_index_from_string(''.join(c for c in end if c.isalpha()))
        end_row = int(''.join(c for c in end if c.isdigit()))

        label_row = start_row
        label_range = f"{get_column_letter(start_col)}{label_row}:{get_column_letter(end_col)}{label_row}"
        value_range = f"{get_column_letter(start_col)}{start_row+1}:{get_column_letter(end_col)}{end_row}"
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)

        lbl = ws[f"{get_column_letter(start_col)}{label_row}"]
        lbl.value = f"  {label}"
        lbl.font = Font(name="Segoe UI", bold=True, color=Colors.TEXT_MUTED, size=8)
        lbl.fill = fill(Colors.BG_SURFACE)
        lbl.alignment = Alignment(horizontal="left", vertical="center")

        val = ws[f"{get_column_letter(start_col)}{start_row+1}"]
        val.value = value_formula
        val.font = Font(name="Segoe UI", bold=True, color=accent_color, size=16)
        val.fill = fill(Colors.BG_SURFACE)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.number_format = number_format

        # Bottom accent line
        for c in range(start_col, end_col + 1):
            bottom_cell = ws.cell(row=end_row, column=c)
            bottom_cell.border = Border(bottom=Side(border_style="thin", color=accent_color))

    def _paint_background(self, ws, rows, cols, color, start_row=1):
        """Paint background only where it doesn't interfere with data rows.
        Skip if start_row > 1 to avoid inflating max_row for data sheets."""
        for r in range(start_row, rows + 1):
            for c in range(1, cols + 1):
                if ws.cell(row=r, column=c).value is None:
                    ws.cell(row=r, column=c).fill = fill(color)

    def _paint_range(self, ws, range_ref, color):
        start, end = range_ref.split(":")
        start_col = openpyxl.utils.column_index_from_string(''.join(c for c in start if c.isalpha()))
        start_row = int(''.join(c for c in start if c.isdigit()))
        end_col = openpyxl.utils.column_index_from_string(''.join(c for c in end if c.isalpha()))
        end_row = int(''.join(c for c in end if c.isdigit()))
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).fill = fill(color)

    def _render_dashboard_charts(self, ws):
        """Render embedded charts that read from other sheets."""
        demo_ws = ws.parent["Demo-Kapital"]
        perf_ws = ws.parent["Performance"]

        # Dynamischen max_row für Equity Curve finden (max 200 ist für Performance genug)
        equity_max_row = 8
        for r in range(8, 500):
            if demo_ws.cell(row=r, column=1).value is None:
                break
            equity_max_row = r
        # Mindestens 2 Punkte für sinnvolle Linie
        equity_max_row = max(equity_max_row, 9)

        # ── Equity Curve (line chart) ──
        equity = LineChart()
        equity.title = None
        equity.legend = None
        equity.style = 2
        equity.height = 8
        equity.width = 24

        # Nur tatsächlich befüllte Zellen referenzieren (titles_from_data=True heißt Zeile 7 ist Header)
        data_ref = Reference(demo_ws, min_col=4, min_row=7, max_row=equity_max_row, max_col=4)
        cats_ref = Reference(demo_ws, min_col=1, min_row=8, max_row=equity_max_row)
        equity.add_data(data_ref, titles_from_data=True)
        equity.set_categories(cats_ref)

        if equity.series:
            s = equity.series[0]
            s.graphicalProperties = GraphicalProperties(
                solidFill=Colors.CYAN.replace("FF", "", 1)
            )
            s.graphicalProperties.line = LineProperties(
                solidFill=Colors.CYAN.replace("FF", "", 1),
                w=28000
            )
            s.smooth = True

        ws.add_chart(equity, "B16")

        # ── Win/Loss Pie ──
        perf_ws["L1"] = "Kategorie"
        perf_ws["M1"] = "Anzahl"
        perf_ws["L2"] = "Gewonnen"
        perf_ws["L3"] = "Verloren"
        perf_ws["M2"] = '=COUNTIF(Trades!J2:J10000,"gewonnen")'
        perf_ws["M3"] = '=COUNTIF(Trades!J2:J10000,"verloren")'

        pie = PieChart()
        pie.title = None
        pie.height = 7
        pie.width = 11

        # Mit titles_from_data=True wird Zeile 1 als Series-Titel genutzt, nicht als Daten
        pie_data = Reference(perf_ws, min_col=13, min_row=1, max_row=3)
        pie_labels = Reference(perf_ws, min_col=12, min_row=2, max_row=3)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        pie.dataLabels = DataLabelList(
            showPercent=True,
            showCatName=True,
            showVal=False,
            showSerName=False,
        )
        ws.add_chart(pie, "B34")

        # ── Asset P&L Bar Chart ──
        perf_ws["L6"] = "Asset"
        perf_ws["M6"] = "P&L"
        assets = ["BTC/USD", "XAU/USD", "US500", "EUR/USD"]
        for i, asset in enumerate(assets):
            perf_ws[f"L{7+i}"] = asset
            perf_ws[f"M{7+i}"] = f'=SUMIFS(Trades!K2:K10000,Trades!C2:C10000,"{asset}")'

        bar = BarChart()
        bar.type = "bar"
        bar.title = None
        bar.legend = None
        bar.height = 7
        bar.width = 13

        bar_data = Reference(perf_ws, min_col=13, min_row=6, max_row=10)
        bar_cats = Reference(perf_ws, min_col=12, min_row=7, max_row=10)
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, "H34")

    # ─── Sheet: Demo-Kapital (improved) ──────────────────────────────────────
    def _create_demo_kapital_sheet(self, wb):
        ws = wb.create_sheet("Demo-Kapital")
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = Colors.CYAN.replace("FF", "", 1)

        # Background nur auf Bereich rund um KPI-Header (Rows 1-7)
        for r in range(1, 8):
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = fill(Colors.BG_DEEP)

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"].value = "  KAPITAL-TRACKING"
        ws["A1"].font = Font(name="Segoe UI", bold=True, color=Colors.GOLD, size=16)
        ws["A1"].fill = fill(Colors.BG_HEADER)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws["A1"].border = bottom_accent_border(Colors.GOLD, "medium")
        ws.row_dimensions[1].height = 36

        # KPI row (rows 3-4)
        kpi_defs = [
            ("STARTKAPITAL",      "B3", "B4", "=1000",                                      "€#,##0.00",                 Colors.CYAN),
            ("AKTUELLES KAPITAL", "D3", "D4", "=B4+SUM(C8:C10000)",                         "€#,##0.00",                 Colors.EMERALD),
            ("GESAMT P&L",        "F3", "F4", "=D4-B4",                                     "€#,##0.00;[Red]-€#,##0.00", Colors.GOLD),
            ("ROI",               "H3", "H4", "=IF(B4>0,(D4-B4)/B4,0)",                     "0.00%;[Red]-0.00%",         Colors.PURPLE),
        ]
        for label, lcell, vcell, formula, fmt, color in kpi_defs:
            ws[lcell].value = label
            ws[lcell].font = font(bold=True, color=Colors.TEXT_MUTED, size=9)
            ws[lcell].fill = fill(Colors.BG_ELEVATED)
            ws[lcell].alignment = center()
            ws.merge_cells(f"{lcell}:{chr(ord(lcell[0])+1)}{lcell[1:]}")

            ws[vcell].value = formula
            ws[vcell].font = Font(name="Segoe UI", bold=True, color=color, size=18)
            ws[vcell].fill = fill(Colors.BG_SURFACE)
            ws[vcell].alignment = center()
            ws[vcell].number_format = fmt
            ws.merge_cells(f"{vcell}:{chr(ord(vcell[0])+1)}{vcell[1:]}")
            ws[vcell].border = bottom_accent_border(color, "medium")

        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 36
        ws.row_dimensions[5].height = 10

        # Kapitalverlauf table
        ws.merge_cells("A6:H6")
        ws["A6"].value = "   KAPITALVERLAUF"
        ws["A6"].font = Font(name="Segoe UI", bold=True, color=Colors.CYAN, size=11)
        ws["A6"].fill = fill(Colors.BG_ELEVATED)
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A6"].border = bottom_accent_border(Colors.CYAN, "medium")
        ws.row_dimensions[6].height = 26

        headers = ["Datum", "P&L Tag", "Kumulativ P&L", "Kapital", "Trades", "Gewonnen", "Verloren", "Win Rate"]
        widths  = [14, 14, 16, 14, 11, 12, 12, 12]
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = w
            cell = ws.cell(row=7, column=i, value=h)
            cell.font = font(bold=True, color=Colors.WHITE, size=10)
            cell.fill = fill(Colors.BG_HEADER)
            cell.alignment = center()
            cell.border = bottom_accent_border(Colors.GOLD, "medium")
        ws.row_dimensions[7].height = 28

        # Number formats for future data rows (apply to whole column)
        ws.column_dimensions["B"].number_format = "€#,##0.00;[Red]-€#,##0.00"
        ws.column_dimensions["C"].number_format = "€#,##0.00;[Red]-€#,##0.00"
        ws.column_dimensions["D"].number_format = "€#,##0.00"
        ws.column_dimensions["H"].number_format = "0.0%"

        # Conditional formatting: data bar on "Kapital" column
        bar_rule = DataBarRule(start_type="min", end_type="max",
                               color=Colors.CYAN.replace("FF", "", 1),
                               showValue=True)
        ws.conditional_formatting.add("D8:D10000", bar_rule)

        # Color scale on P&L Tag
        pnl_rule = ColorScaleRule(
            start_type="min", start_color=Colors.RED.replace("FF", "", 1),
            mid_type="num", mid_value=0, mid_color=Colors.BG_ELEVATED.replace("FF", "", 1),
            end_type="max", end_color=Colors.EMERALD.replace("FF", "", 1),
        )
        ws.conditional_formatting.add("B8:B10000", pnl_rule)
        ws.conditional_formatting.add("C8:C10000", pnl_rule)

    # ─── Sheet: Trades (improved with conditional formatting) ────────────────
    def _create_trades_sheet(self, wb):
        ws = wb.create_sheet("Trades")
        ws.sheet_view.showGridLines = False
        ws.sheet_view.freezePanes = "A2"
        ws.sheet_properties.tabColor = Colors.EMERALD.replace("FF", "", 1)

        # Nur Header-Row Hintergrund (siehe weiter unten)

        headers = ["Datum", "Uhrzeit", "Asset", "Direction", "Einsatz €", "SL %", "TP %",
                   "R:R", "Status", "Ergebnis", "P&L €", "Kapital danach"]
        widths  = [12, 10, 12, 12, 13, 9, 9, 9, 12, 13, 13, 16]
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = font(bold=True, color=Colors.WHITE, size=10)
            cell.fill = fill(Colors.BG_HEADER)
            cell.alignment = center()
            cell.border = bottom_accent_border(Colors.EMERALD, "medium")
        ws.row_dimensions[1].height = 30

        # Column-level number formats
        for col_letter in ["E", "K", "L"]:
            for row in range(2, 10001):
                ws[f"{col_letter}{row}"].number_format = "€#,##0.00;[Red]-€#,##0.00"

        # Conditional formatting on Direction column (D)
        green_font = Font(name="Segoe UI", bold=True, color=Colors.EMERALD)
        red_font   = Font(name="Segoe UI", bold=True, color=Colors.RED)

        ws.conditional_formatting.add(
            "D2:D10000",
            CellIsRule(operator="equal", formula=['"LONG"'],
                       stopIfTrue=False,
                       font=green_font,
                       fill=fill(Colors.BG_SURFACE))
        )
        ws.conditional_formatting.add(
            "D2:D10000",
            CellIsRule(operator="equal", formula=['"SHORT"'],
                       stopIfTrue=False,
                       font=red_font,
                       fill=fill(Colors.BG_SURFACE))
        )

        # Result column (J)
        ws.conditional_formatting.add(
            "J2:J10000",
            CellIsRule(operator="equal", formula=['"gewonnen"'],
                       stopIfTrue=False,
                       font=Font(name="Segoe UI", bold=True, color=Colors.EMERALD),
                       fill=fill(Colors.BG_SURFACE))
        )
        ws.conditional_formatting.add(
            "J2:J10000",
            CellIsRule(operator="equal", formula=['"verloren"'],
                       stopIfTrue=False,
                       font=Font(name="Segoe UI", bold=True, color=Colors.RED),
                       fill=fill(Colors.BG_SURFACE))
        )
        ws.conditional_formatting.add(
            "J2:J10000",
            CellIsRule(operator="equal", formula=['"offen"'],
                       stopIfTrue=False,
                       font=Font(name="Segoe UI", bold=True, color=Colors.AMBER),
                       fill=fill(Colors.BG_SURFACE))
        )

        # P&L color scale on K
        pnl_scale = ColorScaleRule(
            start_type="min", start_color=Colors.RED.replace("FF", "", 1),
            mid_type="num", mid_value=0, mid_color=Colors.BG_ELEVATED.replace("FF", "", 1),
            end_type="max", end_color=Colors.EMERALD.replace("FF", "", 1),
        )
        ws.conditional_formatting.add("K2:K10000", pnl_scale)

        # Zebra striping via formula rule
        stripe_rule = Rule(type="expression",
                           formula=["MOD(ROW(),2)=0"],
                           dxf=DifferentialStyle(fill=fill(Colors.BG_ACCENT)))
        stripe_rule.priority = 100  # lower priority so other rules win
        ws.conditional_formatting.add("A2:L10000", stripe_rule)

    # ─── Sheet: Analyse-Log ──────────────────────────────────────────────────
    def _create_analyse_sheet(self, wb):
        ws = wb.create_sheet("Analyse-Log")
        ws.sheet_view.showGridLines = False
        ws.sheet_view.freezePanes = "A2"
        ws.sheet_properties.tabColor = Colors.CYAN.replace("FF", "", 1)

        headers = ["Datum", "Uhrzeit", "Asset", "Action", "Konfidenz", "SL %", "TP %",
                   "R:R", "Score", "Urgency", "Zusammenfassung"]
        widths  = [12, 10, 12, 11, 12, 9, 9, 9, 11, 12, 65]
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = font(bold=True, color=Colors.WHITE, size=10)
            cell.fill = fill(Colors.BG_HEADER)
            cell.alignment = center()
            cell.border = bottom_accent_border(Colors.CYAN, "medium")
        ws.row_dimensions[1].height = 30

        # Color-code action column
        ws.conditional_formatting.add(
            "D2:D10000",
            CellIsRule(operator="equal", formula=['"BUY"'],
                       font=Font(name="Segoe UI", bold=True, color=Colors.EMERALD),
                       fill=fill(Colors.BG_SURFACE))
        )
        ws.conditional_formatting.add(
            "D2:D10000",
            CellIsRule(operator="equal", formula=['"SELL"'],
                       font=Font(name="Segoe UI", bold=True, color=Colors.RED),
                       fill=fill(Colors.BG_SURFACE))
        )

        # Urgency column
        ws.conditional_formatting.add(
            "J2:J10000",
            CellIsRule(operator="equal", formula=['"immediate"'],
                       font=Font(name="Segoe UI", bold=True, color=Colors.GOLD),
                       fill=fill(Colors.BG_SURFACE))
        )
        ws.conditional_formatting.add(
            "J2:J10000",
            CellIsRule(operator="equal", formula=['"wait"'],
                       font=Font(name="Segoe UI", bold=True, color=Colors.TEXT_SECONDARY),
                       fill=fill(Colors.BG_SURFACE))
        )

        # Zebra striping
        stripe_rule = Rule(type="expression",
                           formula=["MOD(ROW(),2)=0"],
                           dxf=DifferentialStyle(fill=fill(Colors.BG_ACCENT)))
        stripe_rule.priority = 100
        ws.conditional_formatting.add("A2:K10000", stripe_rule)

    # ─── Sheet: Performance ──────────────────────────────────────────────────
    def _create_performance_sheet(self, wb):
        ws = wb.create_sheet("Performance")
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = Colors.PURPLE.replace("FF", "", 1)

        # Background nur auf KPI+Header-Bereich
        for r in range(1, 8):
            for c in range(1, 15):
                ws.cell(row=r, column=c).fill = fill(Colors.BG_DEEP)

        ws.merge_cells("A1:I1")
        ws["A1"].value = "   PERFORMANCE ÜBERSICHT"
        ws["A1"].font = Font(name="Segoe UI", bold=True, color=Colors.PURPLE, size=14)
        ws["A1"].fill = fill(Colors.BG_HEADER)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A1"].border = bottom_accent_border(Colors.PURPLE, "medium")
        ws.row_dimensions[1].height = 36

        # KPI cards
        kpi_defs = [
            ("GESAMT TRADES", "B3", "B4", "=COUNTA(Trades!A2:A10000)",                      "0",                 Colors.CYAN),
            ("WIN RATE",      "D3", "D4", '=IF(COUNTA(Trades!A2:A10000)>0,COUNTIF(Trades!J2:J10000,"gewonnen")/COUNTA(Trades!A2:A10000),0)', "0.0%", Colors.EMERALD),
            ("GESAMT P&L",    "F3", "F4", "=SUM(Trades!K2:K10000)",                         "€#,##0.00;[Red]-€#,##0.00", Colors.GOLD),
            ("BESTER TRADE",  "H3", "H4", "=IFERROR(MAX(Trades!K2:K10000),0)",              "€#,##0.00",                  Colors.EMERALD),
        ]
        for label, lcell, vcell, formula, fmt, color in kpi_defs:
            ws[lcell].value = label
            ws[lcell].font = font(bold=True, color=Colors.TEXT_MUTED, size=9)
            ws[lcell].fill = fill(Colors.BG_ELEVATED)
            ws[lcell].alignment = center()
            ws.merge_cells(f"{lcell}:{chr(ord(lcell[0])+1)}{lcell[1:]}")

            ws[vcell].value = formula
            ws[vcell].font = Font(name="Segoe UI", bold=True, color=color, size=18)
            ws[vcell].fill = fill(Colors.BG_SURFACE)
            ws[vcell].alignment = center()
            ws[vcell].number_format = fmt
            ws.merge_cells(f"{vcell}:{chr(ord(vcell[0])+1)}{vcell[1:]}")
            ws[vcell].border = bottom_accent_border(color, "medium")

        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 34
        ws.row_dimensions[5].height = 10

        # Monatsübersicht table
        ws.merge_cells("A6:I6")
        ws["A6"].value = "   MONATS-ÜBERSICHT"
        ws["A6"].font = Font(name="Segoe UI", bold=True, color=Colors.PURPLE, size=11)
        ws["A6"].fill = fill(Colors.BG_ELEVATED)
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A6"].border = bottom_accent_border(Colors.PURPLE, "medium")
        ws.row_dimensions[6].height = 26

        headers = ["Monat", "Trades", "Gewonnen", "Verloren", "Win Rate", "P&L €", "ROI %", "Bester", "Schlechtester"]
        widths  = [14, 10, 12, 12, 12, 14, 10, 12, 16]
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            cell = ws.cell(row=7, column=i, value=h)
            cell.font = font(bold=True, color=Colors.WHITE, size=10)
            cell.fill = fill(Colors.BG_HEADER)
            cell.alignment = center()
            cell.border = bottom_accent_border(Colors.PURPLE, "medium")
        ws.row_dimensions[7].height = 26

        # Number formats
        for r in range(8, 500):
            ws[f"E{r}"].number_format = "0.0%"
            ws[f"F{r}"].number_format = "€#,##0.00;[Red]-€#,##0.00"
            ws[f"G{r}"].number_format = "0.00%"
            ws[f"H{r}"].number_format = "€#,##0.00;[Red]-€#,##0.00"
            ws[f"I{r}"].number_format = "€#,##0.00;[Red]-€#,##0.00"

    # ─── Sheet: Einstellungen ────────────────────────────────────────────────
    def _create_settings_sheet(self, wb):
        ws = wb.create_sheet("Einstellungen")
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = Colors.AMBER.replace("FF", "", 1)

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 36

        ws.merge_cells("A1:B1")
        ws["A1"].value = "   EINSTELLUNGEN"
        ws["A1"].font = Font(name="Segoe UI", bold=True, color=Colors.AMBER, size=14)
        ws["A1"].fill = fill(Colors.BG_HEADER)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A1"].border = bottom_accent_border(Colors.AMBER, "medium")
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 10

        settings = [
            ("SECTION", "System"),
            ("Version",         "Trading Multi-Agent v3.0"),
            ("Erstellt am",     datetime.now().strftime("%d.%m.%Y %H:%M")),
            ("SECTION", "Trading-Config"),
            ("Assets",          os.getenv("TRADING_ASSETS", "EUR/USD,BTC/USD,XAU/USD,US500")),
            ("Strategie",       os.getenv("TRADING_STRATEGY", "adaptive")),
            ("Max Risiko %",    os.getenv("MAX_RISK_PCT", "2")),
            ("Stop Loss %",     os.getenv("STOP_LOSS_PCT", "1.5")),
            ("Take Profit %",   os.getenv("TAKE_PROFIT_PCT", "3.0")),
            ("Position Size €", os.getenv("POSITION_SIZE_EUR", "1000")),
            ("Auto Trade",      os.getenv("AUTO_TRADE", "false")),
            ("Capital.com",     "DEMO" if os.getenv("CAPITAL_DEMO", "true").lower() == "true" else "LIVE"),
            ("SECTION", "Demo-Kapital"),
            ("Startkapital €",  os.getenv("DEMO_STARTKAPITAL", "1000")),
            ("Min. Konfidenz",  os.getenv("MIN_CONFIDENCE", "70")),
            ("Schedule Stunden", os.getenv("SCHEDULE_INTERVAL_HOURS", "1")),
        ]

        row = 3
        for k, v in settings:
            if k == "SECTION":
                ws.merge_cells(f"A{row}:B{row}")
                c = ws.cell(row=row, column=1, value=f"   {v}")
                c.font = Font(name="Segoe UI", bold=True, color=Colors.GOLD, size=11)
                c.fill = fill(Colors.BG_ELEVATED)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bottom_accent_border(Colors.GOLD, "thin")
                ws.row_dimensions[row].height = 26
                row += 1
                continue

            ck = ws.cell(row=row, column=1, value=k)
            cv = ws.cell(row=row, column=2, value=v)
            ck.font = font(color=Colors.TEXT_SECONDARY, size=10)
            ck.fill = fill(Colors.BG_SURFACE)
            ck.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            cv.font = font(color=Colors.TEXT_PRIMARY, size=10, bold=True)
            cv.fill = fill(Colors.BG_SURFACE)
            cv.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ck.border = thin_border(Colors.BORDER_SOFT)
            cv.border = thin_border(Colors.BORDER_SOFT)
            ws.row_dimensions[row].height = 22
            row += 1

    # ─── Data Writers (with correct P&L calculation) ─────────────────────────
    def _load_wb(self):
        return openpyxl.load_workbook(self.path)

    def save_analysis(self, result: dict):
        try:
            wb = self._load_wb()
            ws = wb["Analyse-Log"]
            # Next free row (start at 2, skip header)
            row = 2
            while ws.cell(row=row, column=1).value is not None:
                row += 1
            now = datetime.now()
            for d in result.get("decisions", []):
                if d.get("action") == "hold":
                    continue
                action = d.get("action", "").upper()

                cells = [
                    now.strftime("%d.%m.%Y"),
                    now.strftime("%H:%M:%S"),
                    d.get("asset", ""),
                    action,
                    f"{d.get('confidence', 0)}%",
                    f"{d.get('stopLoss', 0):.1f}%",
                    f"{d.get('takeProfit', 0):.1f}%",
                    f"{d.get('riskReward', 0):.1f}:1",
                    f"{result.get('sessionScore', 0)}/100",
                    d.get("urgency", ""),
                    d.get("summary", "")[:300],
                ]
                for i, v in enumerate(cells, start=1):
                    cell = ws.cell(row=row, column=i, value=v)
                    cell.font = font(color=Colors.TEXT_PRIMARY, size=10)
                    cell.fill = fill(Colors.BG_SURFACE)
                    cell.alignment = left() if i == 11 else center()
                    cell.border = thin_border(Colors.BORDER_SOFT)
                ws.row_dimensions[row].height = 22
                row += 1
            wb.save(self.path)
        except Exception as e:
            log.error(f"save_analysis Fehler: {e}")

    def save_trade(self, trade: dict):
        """Save trade with CORRECT P&L calculation based on actual outcome."""
        try:
            wb = self._load_wb()
            ws = wb["Trades"]
            row = 2
            while ws.cell(row=row, column=1).value is not None:
                row += 1
            now = datetime.now()

            direction = trade.get("direction", trade.get("action", "")).upper()
            einsatz   = float(trade.get("einsatz", trade.get("size", 0)))
            sl_pct    = float(trade.get("sl_pct", trade.get("stopLoss", 1.5)))
            tp_pct    = float(trade.get("tp_pct", trade.get("takeProfit", 3.0)))
            rr        = float(trade.get("rr", 0))
            status    = trade.get("status", "offen")
            ergebnis  = trade.get("ergebnis", "offen")

            # Correct P&L calculation based on result
            if "pnl" in trade and trade["pnl"] != 0:
                pnl = float(trade["pnl"])
            elif ergebnis == "gewonnen":
                pnl = round(einsatz * tp_pct / 100, 2)
            elif ergebnis == "verloren":
                pnl = round(-einsatz * sl_pct / 100, 2)
            else:
                pnl = 0.0

            kapital_danach = trade.get("kapital_danach", None)

            cells = [
                now.strftime("%d.%m.%Y"),
                now.strftime("%H:%M:%S"),
                trade.get("asset", ""),
                direction,
                einsatz,
                f"{sl_pct:.1f}%",
                f"{tp_pct:.1f}%",
                f"{rr:.1f}:1",
                status,
                ergebnis,
                pnl,
                kapital_danach if kapital_danach is not None else "",
            ]
            for i, v in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.font = font(color=Colors.TEXT_PRIMARY, size=10)
                cell.fill = fill(Colors.BG_SURFACE)
                cell.alignment = center()
                cell.border = thin_border(Colors.BORDER_SOFT)
                # Number formats
                if i in (5, 11, 12):
                    cell.number_format = "€#,##0.00;[Red]-€#,##0.00"
            ws.row_dimensions[row].height = 22
            wb.save(self.path)
        except Exception as e:
            log.error(f"save_trade Fehler: {e}")

    def save_demo_snapshot(self, snapshot: dict):
        try:
            wb = self._load_wb()
            ws = wb["Demo-Kapital"]
            # Richtige nächste Zeile suchen (max_row ist durch Background-Paint verfälscht)
            row = 8
            while ws.cell(row=row, column=1).value is not None:
                row += 1
            cells = [
                snapshot.get("datum", ""),
                snapshot.get("pnl", 0),
                snapshot.get("kumulativ_pnl", 0),
                snapshot.get("kapital", 0),
                snapshot.get("trades", 0),
                snapshot.get("gewonnen", 0),
                snapshot.get("verloren", 0),
                snapshot.get("win_rate", 0) / 100 if snapshot.get("win_rate", 0) > 1 else snapshot.get("win_rate", 0),
            ]
            for i, v in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.font = font(color=Colors.TEXT_PRIMARY, size=10)
                cell.fill = fill(Colors.BG_SURFACE)
                cell.alignment = center()
                cell.border = thin_border(Colors.BORDER_SOFT)
                if i in (2, 3, 4):
                    cell.number_format = "€#,##0.00;[Red]-€#,##0.00"
                elif i == 8:
                    cell.number_format = "0.0%"
            ws.row_dimensions[row].height = 22

            # Equity-Curve Chart auf Dashboard aktualisieren
            self._refresh_equity_chart(wb)

            wb.save(self.path)
        except Exception as e:
            log.error(f"save_demo_snapshot Fehler: {e}")

    def _refresh_equity_chart(self, wb):
        """Aktualisiert den Equity-Curve-Chart auf dem Dashboard nach neuen Snapshots."""
        try:
            dash = wb["Dashboard"]
            demo = wb["Demo-Kapital"]

            # Letzte Zeile mit Daten finden
            max_row = 8
            for r in range(8, 500):
                if demo.cell(row=r, column=1).value is None:
                    break
                max_row = r
            max_row = max(max_row, 9)

            # Alten Equity-Chart entfernen (erster Chart im Dashboard)
            if dash._charts:
                dash._charts = [c for c in dash._charts if not isinstance(c, LineChart)]

            # Neuen Chart erstellen
            equity = LineChart()
            equity.title = None
            equity.legend = None
            equity.style = 2
            equity.height = 8
            equity.width = 24

            data_ref = Reference(demo, min_col=4, min_row=7, max_row=max_row, max_col=4)
            cats_ref = Reference(demo, min_col=1, min_row=8, max_row=max_row)
            equity.add_data(data_ref, titles_from_data=True)
            equity.set_categories(cats_ref)

            if equity.series:
                s = equity.series[0]
                s.graphicalProperties = GraphicalProperties(
                    solidFill=Colors.CYAN.replace("FF", "", 1)
                )
                s.graphicalProperties.line = LineProperties(
                    solidFill=Colors.CYAN.replace("FF", "", 1),
                    w=28000
                )
                s.smooth = True

            dash.add_chart(equity, "B16")
        except Exception as e:
            log.warning(f"_refresh_equity_chart: {e}")

    def get_trade_history(self) -> dict:
        try:
            wb = self._load_wb()
            ws = wb["Trades"]
            trades = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(c is not None for c in row):
                    trades.append({
                        "date": row[0], "time": row[1], "asset": row[2],
                        "direction": row[3], "size": row[4],
                        "sl": row[5], "tp": row[6],
                        "status": row[8], "ergebnis": row[9], "pnl": row[10],
                    })
            return {"trades": trades, "count": len(trades)}
        except Exception as e:
            log.error(f"get_trade_history Fehler: {e}")
            return {"trades": [], "count": 0}
