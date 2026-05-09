"""
demo_tracker.py (v3 - openpyxl only, NO pandas)
────────────────────────────────────────────────
Liest/schreibt Demo-Kapital direkt aus Excel mit openpyxl.
Keine pandas-Dependency mehr - schnellerer Start, schlankerer Container.
"""

import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook

log = logging.getLogger(__name__)

# ─── Konfiguration ──────────────────────────────────────────
DATA_DIR       = os.getenv("DATA_DIR", "/app/data")
STARTKAPITAL   = float(os.getenv("DEMO_STARTKAPITAL", "1000"))
RISIKO_PROZENT = float(os.getenv("MAX_RISK_PCT", "5"))
SL_PROZENT     = float(os.getenv("STOP_LOSS_PCT", "1.0"))
TP_PROZENT     = float(os.getenv("TAKE_PROFIT_PCT", "2.0"))
EXCEL_FILE     = Path(DATA_DIR) / "Trading_Tracker.xlsx"
SHEET_NAME     = "Demo-Kapital"

# Spalten-Reihenfolge (KONSISTENT halten!)
COLUMNS = [
    "Datum", "Uhrzeit", "ID", "Asset", "Action", "Richtung",
    "Konfidenz", "Einsatz", "SL %", "TP %", "SL Absolut", "TP Absolut",
    "R:R", "Entry-Price", "Aktuell", "P&L", "Status",
    "Geöffnet am", "Geschlossen am", "Zusammenfassung", "Score", "Strategie",
]

os.makedirs(DATA_DIR, exist_ok=True)


# ─── Helper-Funktionen ──────────────────────────────────────

def _ensure_workbook():
    """Stellt sicher dass Workbook und Sheet existieren."""
    if not EXCEL_FILE.exists():
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
        log.info(f"✅ Excel erstellt: {EXCEL_FILE}")
        return

    wb = load_workbook(EXCEL_FILE)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
        log.info(f"✅ Sheet '{SHEET_NAME}' erstellt")


def _lade_trades() -> list:
    """Lädt alle Trades als Liste von Dicts."""
    _ensure_workbook()

    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = list(rows[0])
        trades = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            trade = {h: v for h, v in zip(headers, row)}
            trades.append(trade)

        return trades
    except Exception as e:
        log.error(f"Excel lesen Fehler: {e}")
        return []


def _speichere_trades(trades: list):
    """Speichert alle Trades zurück ins Excel."""
    if not trades:
        log.warning("Leere Trade-Liste - überspringe Speichern")
        return

    try:
        if EXCEL_FILE.exists():
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Sheet komplett neu schreiben (cleaner als rows einzeln updaten)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)

        ws.append(COLUMNS)
        for trade in trades:
            row = [trade.get(col, "") for col in COLUMNS]
            ws.append(row)

        wb.save(EXCEL_FILE)
        log.info(f"✅ Excel gespeichert: {len(trades)} Trades")
    except Exception as e:
        log.error(f"Excel speichern Fehler: {e}")


def _validiere_prozent(wert, default: float, max_wert: float = 20.0) -> float:
    """Stellt sicher dass ein Wert ein gültiger Prozentsatz ist."""
    try:
        wert = float(wert)
        if wert <= 0 or wert > max_wert:
            return default
        return wert
    except Exception:
        return default


def _safe_float(value, default: float = 0.0) -> float:
    """Konvertiert sicher zu float."""
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


# ─── Public API ─────────────────────────────────────────────

def signal_oeffnen(signal: dict) -> dict:
    """Öffnet einen neuen Demo-Trade und speichert in Excel."""
    trades = _lade_trades()

    stats = get_statistik()
    kapital = stats["aktuelles_kapital"]
    einsatz = round(kapital * RISIKO_PROZENT / 100, 2)
    einsatz = max(1.0, min(einsatz, kapital))

    sl_pct = _validiere_prozent(signal.get("stopLoss"), SL_PROZENT)
    tp_pct = _validiere_prozent(signal.get("takeProfit"), TP_PROZENT)

    sl_absolut = round(einsatz * sl_pct / 100, 2)
    tp_absolut = round(einsatz * tp_pct / 100, 2)

    trade_id = f"T{len(trades) + 1:04d}"

    neue_zeile = {
        "Datum": datetime.now().strftime("%d.%m.%Y"),
        "Uhrzeit": datetime.now().strftime("%H:%M:%S"),
        "ID": trade_id,
        "Asset": signal.get("asset", ""),
        "Action": signal.get("action", "").upper(),
        "Richtung": signal.get("direction", "").upper(),
        "Konfidenz": int(signal.get("confidence", 0)),
        "Einsatz": einsatz,
        "SL %": sl_pct,
        "TP %": tp_pct,
        "SL Absolut": sl_absolut,
        "TP Absolut": tp_absolut,
        "R:R": round(tp_pct / sl_pct, 2) if sl_pct > 0 else 0,
        "Entry-Price": _safe_float(signal.get("entry_price", 0)),
        "Aktuell": 0.0,
        "P&L": 0.0,
        "Status": "offen",
        "Geöffnet am": datetime.now().isoformat(),
        "Geschlossen am": "",
        "Zusammenfassung": signal.get("summary", ""),
        "Score": signal.get("sessionScore", 0),
        "Strategie": signal.get("strategyUsed", ""),
    }

    trades.append(neue_zeile)
    _speichere_trades(trades)

    log.info(
        f"✅ Demo-Trade: {trade_id} | {neue_zeile['Asset']} {neue_zeile['Action']} | "
        f"€{einsatz} | SL:{sl_pct}% TP:{tp_pct}% | Entry:{neue_zeile['Entry-Price']}"
    )

    return neue_zeile


def trade_schliessen(trade_id: str, ergebnis: str, pnl_override: Optional[float] = None) -> dict:
    """Schließt einen offenen Demo-Trade."""
    trades = _lade_trades()

    if not trades:
        log.warning(f"Trade {trade_id} nicht gefunden (Excel leer)")
        return {}

    idx = None
    for i, t in enumerate(trades):
        if t.get("ID") == trade_id:
            idx = i
            break

    if idx is None:
        log.warning(f"Trade {trade_id} nicht gefunden")
        return {}

    trade = trades[idx]

    if trade.get("Status") != "offen":
        log.warning(f"Trade {trade_id} ist nicht offen")
        return trade

    if pnl_override is not None:
        pnl = pnl_override
    elif ergebnis == "gewonnen":
        pnl = _safe_float(trade.get("TP Absolut", 0))
    elif ergebnis == "verloren":
        pnl = -_safe_float(trade.get("SL Absolut", 0))
    else:
        pnl = 0.0

    trades[idx]["Status"] = ergebnis
    trades[idx]["P&L"] = round(pnl, 2)
    trades[idx]["Geschlossen am"] = datetime.now().isoformat()

    _speichere_trades(trades)

    log.info(
        f"{'✅' if ergebnis == 'gewonnen' else '❌'} Trade {trade_id} | "
        f"{ergebnis.upper()} | P&L: {'+' if pnl >= 0 else ''}€{pnl:.2f}"
    )

    return trades[idx]


def get_offene_trades() -> list:
    """Gibt alle offenen Trades zurück."""
    trades = _lade_trades()
    return [t for t in trades if t.get("Status") == "offen"]


def get_statistik() -> dict:
    """Liest Statistik direkt aus Excel."""
    trades = _lade_trades()

    if not trades:
        return {
            "startkapital": STARTKAPITAL,
            "aktuelles_kapital": STARTKAPITAL,
            "pnl_gesamt": 0.0,
            "erstellt_am": datetime.now().isoformat(),
            "statistik": {
                "gesamt_trades": 0,
                "gewonnen": 0,
                "verloren": 0,
                "offen": 0,
                "gesamt_pnl": 0.0,
                "beste_trade": 0.0,
                "schlechtester_trade": 0.0,
                "win_rate": 0.0,
                "roi": 0.0,
                "max_drawdown": 0.0,
            },
            "tages_snapshots": [],
            "offene_trades": [],
            "letzte_trades": [],
        }

    offene = [t for t in trades if t.get("Status") == "offen"]
    geschlossene = [t for t in trades if t.get("Status") in ("gewonnen", "verloren", "breakeven")]
    gewonnen = sum(1 for t in trades if t.get("Status") == "gewonnen")
    verloren = sum(1 for t in trades if t.get("Status") == "verloren")

    gesamt_pnl = round(sum(_safe_float(t.get("P&L", 0)) for t in trades), 2)
    aktuelles_kapital = round(STARTKAPITAL + gesamt_pnl, 2)

    abgeschlossen = gewonnen + verloren
    win_rate = round(gewonnen / abgeschlossen * 100, 1) if abgeschlossen > 0 else 0.0

    roi = round((aktuelles_kapital - STARTKAPITAL) / STARTKAPITAL * 100, 2) if STARTKAPITAL > 0 else 0.0

    beste_trade = 0.0
    schlechtester_trade = 0.0
    if geschlossene:
        pnls = [_safe_float(t.get("P&L", 0)) for t in geschlossene]
        beste_trade = round(max(pnls), 2)
        schlechtester_trade = round(min(pnls), 2)

    # Max Drawdown
    max_drawdown = 0.0
    kapital_progression = [STARTKAPITAL]
    for t in trades:
        kapital_progression.append(kapital_progression[-1] + _safe_float(t.get("P&L", 0)))
    peak = max(kapital_progression)
    for k in kapital_progression:
        drawdown = (peak - k) / peak * 100 if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    max_drawdown = round(max_drawdown, 2)

    letzte = sorted(
        [t for t in trades if t.get("Status") != "offen"],
        key=lambda t: str(t.get("Geschlossen am", "")),
        reverse=True
    )[:20]

    return {
        "startkapital": STARTKAPITAL,
        "aktuelles_kapital": aktuelles_kapital,
        "pnl_gesamt": gesamt_pnl,
        "erstellt_am": datetime.now().isoformat(),
        "statistik": {
            "gesamt_trades": len(trades),
            "gewonnen": gewonnen,
            "verloren": verloren,
            "offen": len(offene),
            "gesamt_pnl": gesamt_pnl,
            "beste_trade": beste_trade,
            "schlechtester_trade": schlechtester_trade,
            "win_rate": win_rate,
            "roi": roi,
            "max_drawdown": max_drawdown,
        },
        "tages_snapshots": [],
        "offene_trades": offene,
        "letzte_trades": letzte,
    }


def tages_snapshot():
    """Speichert täglichen Kapital-Snapshot."""
    stats = get_statistik()
    heute = datetime.now().strftime("%d.%m.%Y")
    log.info(f"📊 Snapshot: {heute} | €{stats['aktuelles_kapital']:.2f}")


def generiere_tages_report() -> str:
    """Generiert Tagesreport aus Excel-Daten."""
    stats = get_statistik()
    kapital = stats["aktuelles_kapital"]
    start = stats["startkapital"]
    pnl = stats["pnl_gesamt"]
    roi = stats["statistik"]["roi"]
    wr = stats["statistik"]["win_rate"]
    offen = stats["statistik"]["offen"]
    gewon = stats["statistik"]["gewonnen"]
    verl = stats["statistik"]["verloren"]

    return (
        f"📊 *TRADING DEMO - TAGESREPORT*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 Kapital: *€{kapital:.2f}*\n"
        f"{'📈' if pnl >= 0 else '📉'} Gesamt P&L: *{'+' if pnl >= 0 else ''}€{pnl:.2f}* "
        f"({'+' if roi >= 0 else ''}{roi:.1f}%)\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🎯 Win Rate: *{wr:.1f}%*\n"
        f"✅ Gewonnen: *{gewon}* | ❌ Verloren: *{verl}*\n"
        f"🔄 Offen: *{offen}*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📈 Startkapital: €{start:.2f}\n"
        f"🤖 _Trading Multi-Agent v3.0_"
    )
